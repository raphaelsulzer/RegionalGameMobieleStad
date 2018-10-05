# -*- coding: utf-8 -*-
"""
/***************************************************************************
 PlanningToolClassDockWidget
                                 A QGIS plugin
 The Storm Help plugin provides help in case of a storm
                             -------------------
        begin                : 2017-01-07
        git sha              : $Format:%H$
        copyright            : (C) 2017 by Raphael Sulzer
        email                : raphaelsulzer@gmx.de
 ***************************************************************************/

/***************************************************************************
 *                                                                         *
 *   This program is free software; you can redistribute it and/or modify  *
 *   it under the terms of the GNU General Public License as published by  *
 *   the Free Software Foundation; either version 2 of the License, or     *
 *   (at your option) any later version.                                   *
 *                                                                         *
 ***************************************************************************/
"""

import os
import math

from PyQt4 import QtGui, QtCore, uic
from qgis.core import *
from qgis.core import QgsGeometry, QgsMapLayerRegistry
# from PyQt4 import QtCore
# from PyQt4 import QtGui

from qgis.gui import *
from qgis.gui import QgsMapTool
from qgis.networkanalysis import *

from PyQt4.QtGui import QCursor, QPixmap
from PyQt4.QtCore import Qt, pyqtSignal, QPoint
#from PyQt4.QtCore import pyqtSignal
from . import utility_functions as uf

# this is for adding the "external" folder to the system path, because the QGIS python is only looking in the system path for python packages
# you can see that this works by doing import sys, sys.path inside the QGIS python console
import os, sys
sys.path.append(os.path.join(os.path.dirname(__file__), "external"))

import processing
import random
import numpy as np
import openpyxl
import xlwings as xw
# matplotlib for the charts
from matplotlib.backends.backend_qt4agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure





FORM_BASE, _ = uic.loadUiType(os.path.join(
    os.path.dirname(__file__), 'PlanningTool_dockwidget_base.ui'))

FORM_INDICATORS, _ = uic.loadUiType(os.path.join(
    os.path.dirname(__file__), 'PlanningTool_chart.ui'))

FORM_HOUSING, _ = uic.loadUiType(os.path.join(
    os.path.dirname(__file__), 'PlanningTool_housing.ui'))

FORM_INFRASTRUCTURE, _ = uic.loadUiType(os.path.join(
    os.path.dirname(__file__), 'PlanningTool_infrastructure.ui'))

class HousingInput(QtGui.QDialog, FORM_HOUSING, QgsMapTool):
    def __init__(self, iface, parent=None, chart=None):

        super(HousingInput, self).__init__(parent)

        # Set up the user interface from Designer.
        # After setupUI you can access any designer object by doing
        # self.<objectname>, and you can use autoconnect slots - see
        # http://qt-project.org/doc/qt-4.8/designer-using-a-ui-file.html
        # #widgets-and-dialogs-with-auto-connect
        self.setupUi(self)

        self.iface = iface

        self.canvas = self.iface.mapCanvas()


        #signal slot for closing indicator window
        self.okHousing.clicked.connect(self.submitHousingInput)
        self.closeHousing.clicked.connect(self.closeHousingInput)

        # # generate canvas clicked signal here
        # self.emitPoint = QgsMapToolEmitPoint(self.canvas)
        # self.emitPoint.canvasClicked.connect(self.selectNearestFeature)
        #
        # # activate coordinate capture tool
        # self.canvas.setMapTool(self.emitPoint)


    def submitHousingInput(self):
        # save in excel sheet
        self.hide()

    # def selectNearestFeature(self, mouseEvent):
    #     """
    #     Each time the mouse is clicked on the map canvas, perform
    #     the following tasks:
    #         Loop through all visible vector layers and for each:
    #             Ensure no features are selected
    #             Determine the distance of the closes feature in the layer to the mouse click
    #             Keep track of the layer id and id of the closest feature
    #         Select the id of the closes feature
    #     """
    #
    #     layerData = []
    #
    #     for layer in self.canvas.layers():
    #
    #         if layer.type() != QgsMapLayer.VectorLayer:
    #             # Ignore this layer as it's not a vector
    #             continue
    #
    #         if layer.featureCount() == 0:
    #             # There are no features - skip
    #             continue
    #
    #         layer.removeSelection()
    #
    #         # print(type(mouseEvent.pos()))
    #         # print(type(layer))
    #
    #         # Determine the location of the click in real-world coords
    #         #layerPoint = self.toLayerCoordinates(layer, mouseEvent.pos())
    #         mapPoint = self.toMapCoordinates(mouseEvent.pos())
    #         canvasPoint = self.toCanvasCoordinates(mapPoint)
    #         layerPoint = self.toLayerCoordinates(layer, mapPoint)
    #
    #         print('mapPoint:', mapPoint)
    #         print('canvasPoint:', canvasPoint)
    #         print('layerPoint:', layerPoint)
    #
    #         shortestDistance = float("inf")
    #         closestFeatureId = -1
    #
    #         # Loop through all features in the layer
    #         for f in layer.getFeatures():
    #             dist = f.geometry().distance(QgsGeometry.fromPoint(layerPoint))
    #             #print(f.id())
    #             #print(dist)
    #             if dist < shortestDistance:
    #                 shortestDistance = dist
    #                 closestFeatureId = f.id()
    #
    #         info = (layer, closestFeatureId, shortestDistance)
    #         layerData.append(info)
    #
    #
    #
    #     if not len(layerData) > 0:
    #         # Looks like no vector layers were found - do nothing
    #         return
    #
    #     # Sort the layer information by shortest distance
    #     layerData.sort(key=lambda element: element[2], reverse=False)
    #
    #     # Select the closest feature
    #     layerWithClosestFeature, closestFeatureId, shortestDistance = layerData[0]
    #     layerWithClosestFeature.select(closestFeatureId)

    def closeHousingInput(self):
        self.hide()


class InfrastructureInput(QtGui.QDialog, FORM_INFRASTRUCTURE):
    def __init__(self, iface, parent=None, chart=None):

        super(InfrastructureInput, self).__init__(parent)

        # Set up the user interface from Designer.
        # After setupUI you can access any designer object by doing
        # self.<objectname>, and you can use autoconnect slots - see
        # http://qt-project.org/doc/qt-4.8/designer-using-a-ui-file.html
        # #widgets-and-dialogs-with-auto-connect
        self.setupUi(self)

        self.iface = iface

        self.chart = chart

        #signal slot for closing indicator window
        # TODO: closeInfrastructure is actually cancel infrastructure, which should maybe actually restore the values that were
        # their, before the input windows was opened
        self.closeInfrastructure.clicked.connect(self.closeInfrastructureInput)
        self.okInfrastructure.clicked.connect(self.saveValue)


        # get project ID and corresponding data from excel sheet
        self.layer = self.iface.activeLayer()
        # self.id is the row number (zero starting) of the attributes table in QGIS, so not the actual id column value
        # thas where the +4 in the getValue call below is coming from
        temp1, temp2 = uf.getFieldValues(self.layer, 'id', null=False, selection=True)
        # guard for when not exactly one project is selected
        if len(temp1) != 1:
            return
        self.id = temp1[0]
        temp1 = None
        temp2 = None
        #print "die project ID aus QGIS: ", self.id

        # get excel row by id of project id from QGIS
        self.excel_file = os.path.join(os.path.dirname(__file__), 'data', 'excel_data.xlsm')
        self.sheet_name = 'INPUT - Infra Projects'
        excel_id = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='A', row_start=self.id)[0]
        #print "excel id: ", excel_id

        # get values with openpyxl
        iC = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='P', row_start=self.id)[0]
        iA = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='Q', row_start=self.id)[0]
        iE = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='R', row_start=self.id)[0]
        iP = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='S', row_start=self.id)[0]
        iZ = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='T', row_start=self.id)[0]
        iH = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='U', row_start=self.id)[0]
        iM = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='V', row_start=self.id)[0]
        # TODO replace these getValue calls with getValue_xlwings calls, because they always fuck up the excel file!!
        #   should also be faster, if that for some reason doesn't work, I could also try to fix the getValue function, e.g. by using a in_memory_file


        # # get values with xlwings
        # iC = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='P', row_start=self.id)[0]
        # iA = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='Q', row_start=self.id)[0]
        # iE = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='R', row_start=self.id)[0]
        # iP = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='S', row_start=self.id)[0]
        # iZ = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='T', row_start=self.id)[0]
        # iH = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='U', row_start=self.id)[0]
        # iM = self.getValue(filepath=self.excel_file, sheetname=self.sheet_name, col='V', row_start=self.id)[0]


        if iC == 1:
            self.inputYes.setChecked(True)
        else:
            self.inputNo.setChecked(True)

        self.inputAmsterdam.setText(str(iA))
        self.inputEdam.setText(str(iE))
        self.inputHoorn.setText(str(iP))
        self.inputPurmerend.setText(str(iZ))
        self.inputZaanstad.setText(str(iH))
        self.inputProvince.setText(str(iM))



    # save value
    def saveValue(self):

        iC = self.inputYes.isChecked()
        iC2 = self.inputNo.isChecked()
        iA = self.inputAmsterdam.text()
        iE = self.inputEdam.text()
        iH = self.inputHoorn.text()
        iP = self.inputPurmerend.text()
        iZ = self.inputZaanstad.text()
        iPr = self.inputProvince.text()
        iM = self.inputMinistry.text()


        #take from input field and save to excel file, depending on polygonID = row in excel file
        #column depending on which input field


        srcfile = openpyxl.load_workbook(self.excel_file, read_only=False, keep_vba=True)  # to open the excel sheet and if it has macros
        sheet = srcfile.get_sheet_by_name(self.sheet_name)  # get sheetname from the file

        # project id is at row+2 in excel, thats why I need to introduce this skip variable
        k = 2
        if iC == True and iC2 == False:
            sheet['P' + str(self.id+k)] = 1
        elif iC == False and iC2 == True:
            sheet['P' + str(self.id+k)] = 0
        else:
            # chase case where both yes and no are checked
            uf.showMessage(self.iface, 'Please select either "yes" or "no"', type='Info', lev=1, dur=5)
            return
        sheet['Q'+str(self.id+k)] = float(iA)
        sheet['R'+str(self.id+k)] = float(iE)
        sheet['S'+str(self.id+k)] = float(iH)
        sheet['T'+str(self.id+k)] = float(iP)
        sheet['U'+str(self.id+k)] = float(iZ)
        sheet['V'+str(self.id+k)] = float(iPr)
        sheet['W'+str(self.id+k)] = float(iM)

        srcfile.save(self.excel_file)
        print 'new values saved'

        self.chart.refreshPlot()


        # close input window
        #self.hide()




    def getValue(self, filepath=None, sheetname=None, col=None, row_start=None, row_end=None):
        # project id is at row+2 in excel, thats why I need to introduce this skip variable
        k = 2
        row_start = row_start+k
        if row_end == None:
            row_end = row_start
        source_file = openpyxl.load_workbook(filepath, read_only=False, keep_vba=True)  # to open the excel sheet and if it has macros
        sheet = source_file.get_sheet_by_name(sheetname)
        data = []
        for i in range(row_start, row_end + 1):
            val = float(sheet[col + str(i)].value)
            data.append(val)

        source_file.save(self.excel_file)
        return np.array(data)


    def closeInfrastructureInput(self):
        self.hide()


# class IndicatorsChart(QtGui.QDialog, FORM_INDICATORS):
#     pass



class IndicatorsChartDocked(QtGui.QDockWidget, FORM_BASE):


    def __init__(self, iface, parent=None, app=None):

        super(IndicatorsChartDocked, self).__init__(parent)

        # Set up the user interface from Designer.
        # After setupUI you can access any designer object by doing
        # self.<objectname>, and you can use autoconnect slots - see
        # http://qt-project.org/doc/qt-4.8/designer-using-a-ui-file.html
        # #widgets-and-dialogs-with-auto-connect
        self.setupUi(self)

        self.iface = iface

        self.app = app
        self.excel_file = os.path.join(os.path.dirname(__file__), 'data', 'excel_data.xlsm')

        #signal slot for closing indicator window
        #self.closeIndicators.clicked.connect(self.closeIndicatorsChart)


        self.refreshPlot()

    def getValue_xlwings(self, app, filename, sheet, cells):
        # # # run Excel with xlwings
        # # # this does not work on windows yet, but the only reason is probably that it cannot access the sheet because Excel opens
        # # # with a put in Product Key Window. So maybe it is actually better to use win32com.client solution as this seems to work better here
        #
        # TODO: open app and book in planning_tool.py and pass to the dockwidget, same way as done for the plot dockwidget that is passed to infrastructure input class
        # app = xw.App(visible=False)
        book = app.books.open(filename)
        # #book = xw.Book(filename)
        # #app = xw.apps.active
        # app.visible = False

        ##get new value based on UDF (user defined function)
        vals = []
        for cell in cells:
            vals.append(book.sheets[sheet].range(cell).value)

        book.close()
        # app.kill()

        # return value
        return np.asarray(vals)
        #return [-1035, -1907, -3106, -7902, -3487]


    def refreshPlot(self):
        ### INDICATORS

        #self.excel_file = os.path.join(os.path.dirname(__file__), 'data', 'excel_data.xlsm')


        print "refreshing plot"

        ### ACCESIBILITY
        # sheet_name = 'INPUT - Infra Projects'
        # ## Transit accesibility
        # c = []
        # af = self.getValue(filepath=self.excel_file, sheetname=sheet_name, col='AF', row_start=3, row_end=40)
        # aj = self.getValue(filepath=self.excel_file, sheetname=sheet_name, col='AJ', row_start=3, row_end=40)
        # ak = aj * 0.01
        # p = self.getValue(filepath=self.excel_file, sheetname=sheet_name, col='P', row_start=3, row_end=40)
        # ar = p * af * ak
        # c.append(sum(ar))
        # # c4
        # ag = self.getValue(filepath=self.excel_file, sheetname=sheet_name, col='AG', row_start=3, row_end=40)
        # as1 = p * ag * ak
        # c.append(sum(as1))
        # # c5
        # ah = self.getValue(filepath=self.excel_file, sheetname=sheet_name, col='AH', row_start=3, row_end=40)
        # at = p * ah * ak
        # c.append(sum(at))
        # # c6
        # ai = self.getValue(filepath=self.excel_file, sheetname=sheet_name, col='AI', row_start=3, row_end=40)
        # au = p * ai * ak
        # c.append(sum(au))
        # ## Car accesibility
        # d = []
        # # d3
        # aa = self.getValue(filepath=self.excel_file, sheetname=sheet_name, col='AA', row_start=3, row_end=40)
        # aj = self.getValue(filepath=self.excel_file, sheetname=sheet_name, col='AJ', row_start=3, row_end=40)
        # ak = aj * 0.01
        # p = self.getValue(filepath=self.excel_file, sheetname=sheet_name, col='P', row_start=3, row_end=40)
        # aw = p * aa * ak
        # d.append(sum(aw))
        # # d4
        # ab = self.getValue(filepath=self.excel_file, sheetname=sheet_name, col='AB', row_start=3, row_end=40)
        # ax = p * ab * ak
        # d.append(sum(ax))
        # # d5
        # ac = self.getValue(filepath=self.excel_file, sheetname=sheet_name, col='AC', row_start=3, row_end=40)
        # ay = p * ac * ak
        # d.append(sum(ay))
        # # d6
        # ad = self.getValue(filepath=self.excel_file, sheetname=sheet_name, col='AD', row_start=3, row_end=40)
        # az = p * ad * ak
        # d.append(sum(az))
        #
        # cd = np.add(c, d)
        # accesibility = np.append(cd, np.mean(cd))

        ### Accesibility
        accesibility = self.getValue_xlwings(self.app, self.excel_file, 'Indicator 1 Accessibility', ['E3', 'E4', 'E5', 'E6', 'E7'])
        accesibility = np.append(accesibility, np.mean(accesibility))

        ### Market Balance
        market_balance = np.asarray([-1035, -1907, -3106, -7902, -3487])
        #market_balance = self.getValue_xlwings(self.excel_file, 'Indicator 2 Market balance', ['E3', 'E4', 'E5', 'E6', 'E7'])

        ### Market Balance
        finances = np.asarray([-5.01, 0, -35.2, 0, 0, 0])
        #finances = self.getValue_xlwings(self.excel_file, 'Indicator 3 Finances', ['E3', 'E4', 'E5', 'E6', 'E7', 'E8'])



        ### PLOT
        # first clear the Figure() from the chartView layout,
        # so a new one will be printed when function is run several times
        for i in reversed(range(self.chartView.count())):
            self.chartView.itemAt(i).widget().setParent(None)

        # add matplotlib Figure to chartFrame / chartView layout
        self.chart_figure = Figure()
        #self.chart_figure.suptitle("Indicators \n\n ", fontsize=18, fontweight='bold')
        self.chart_canvas = FigureCanvas(self.chart_figure)
        self.chartView.addWidget(self.chart_canvas)
        # plot the subplots
        self.plotChart(self.chart_figure.add_subplot(311), accesibility, "Accessibility", 'b')
        self.plotChart(self.chart_figure.add_subplot(312), market_balance, "Market Balance", 'g')
        self.plotChart(self.chart_figure.add_subplot(313), finances, "Finances", 'r')
        #self.chart_figure.tight_layout()
        # you can actually probably adjust it perfectly with this
        self.chart_figure.tight_layout(rect=[0.1, -0.05, 0.94, 1])
        self.chart_figure.subplots_adjust(hspace=0.53)
        # make background of plot transparent
        self.chart_figure.patch.set_alpha(0.0)
        #self.chart_figure.patch.set_facecolor('red')

        return

    # def getValue(self, filepath=None, sheetname=None, col=None, row_start=0, row_end=1):
    #     source_file = openpyxl.load_workbook(filepath, read_only=False, keep_vba=True)  # to open the excel sheet and if it has macros
    #     sheet = source_file.get_sheet_by_name(sheetname)
    #     data = []
    #     for i in range(row_start, row_end + 1):
    #         val = float(sheet[col + str(i)].value)
    #         data.append(val)
    #     source_file.save(self.excel_file)
    #     return np.array(data)


    def plotChart(self, ax, indicator, indicator_name, color):

        ax.cla()

        N = len(indicator)
        ind = np.arange(N)  # the x locations for the groups
        width = 0.35  # the width of the bars

        #rects1 = ax.bar(ind - width, first, width, color='b')
        rects = ax.bar(ind, indicator, width, color=color, align='center')
        #rects3 = ax.bar(ind + width, first, width, color='r')


        # add some text for labels, title and axes ticks
        ax.set_ylabel('Indicator score')
        #ax.set_xlabel('Region')
        ax.set_xticks(ind)
        if N == 5:
            ax.set_xticklabels(('Edam-Vol.', 'Hoorn', 'Purmerend', 'Zaanstad', 'Province'), rotation=30)
        else:
            ax.set_xticklabels(('Edam-Vol.', 'Hoorn', 'Purmerend', 'Zaanstad', 'Province', 'Ministry'), rotation=30)
        ax.legend((rects[0],), (indicator_name,), fontsize=12)

        #low = min(rects1)

        #ax.set_aspect(aspect='equal')

        if np.mean(indicator) < 0:
            ax.invert_yaxis()
            ax.set_ylim((0, min(indicator)*1.8))
        else:
            ax.set_ylim((0, max(indicator)*1.8))

        def autolabel(rectangle):
            """
            Attach a text label above each bar displaying its height
            """
            for rect in rectangle:
                height = rect.get_height()
                ax.text(rect.get_x() + rect.get_width() / 2., 1.05 * height,
                        '%.3f' % float(height),
                        ha='center', va='bottom')

        autolabel(rects)


    def closeIndicatorsChart(self):
        self.hide()

























